"""
Builds HR_Attrition_Analysis.xlsx with:
- README, Raw_Data (sample), Cleaning_Log, Clean_Data (full)
- KPI_Summary: live formulas (SUMIFS/COUNTIFS/AVERAGEIFS) referencing Clean_Data
- Pivot_Summary: formula-driven cross-tab (Department x Attrition Rate) + chart
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

raw = pd.read_csv('/home/claude/hr-attrition/data/hr_attrition_raw.csv')
clean = pd.read_csv('/home/claude/hr-attrition/data/hr_attrition_clean.csv')

HEADER_FILL = PatternFill('solid', start_color='7030A0')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
TITLE_FONT = Font(bold=True, color='7030A0', name='Arial', size=14)
BODY_FONT = Font(name='Arial', size=10)
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

ws = wb.active
ws.title = 'README'
ws['A1'] = 'HR Employee Attrition Analytics — Excel Workbook'
ws['A1'].font = TITLE_FONT
notes = [
    '',
    'Sheet guide:',
    '1. Raw_Data       — sample of original messy export (full file: data/hr_attrition_raw.csv)',
    '2. Cleaning_Log    — issues found and how each was fixed',
    '3. Clean_Data      — full cleaned dataset (1,470 employees), used for Pivot/Summary',
    '4. Pivot_Summary   — formula-driven cross-tab (Department x Attrition Rate) + chart',
    '5. KPI_Summary     — key HR metrics computed with live formulas',
    '',
    'All KPI formulas reference Clean_Data directly (SUMIFS/COUNTIFS/AVERAGEIFS) —',
    'change any value in Clean_Data and the KPIs recalculate automatically.'
]
for i, line in enumerate(notes, start=2):
    ws[f'A{i}'] = line
    ws[f'A{i}'].font = BODY_FONT
ws.column_dimensions['A'].width = 95

def style_header(sheet, n_cols, row=1):
    for c in range(1, n_cols + 1):
        cell = sheet.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def write_df(sheet, df, start_row=1):
    for j, c in enumerate(df.columns, start=1):
        sheet.cell(row=start_row, column=j, value=c)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            cell = sheet.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
    style_header(sheet, len(df.columns), row=start_row)
    for j, col in enumerate(df.columns, start=1):
        maxlen = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 10)
        sheet.column_dimensions[get_column_letter(j)].width = min(max(maxlen + 2, 10), 26)

# ---------------- Raw_Data (sample) ----------------
ws_raw = wb.create_sheet('Raw_Data')
write_df(ws_raw, raw.head(300))
ws_raw.freeze_panes = 'A2'

# ---------------- Cleaning_Log ----------------
ws_log = wb.create_sheet('Cleaning_Log')
log_data = pd.DataFrame([
    ['Duplicate rows', '10 exact duplicate employee rows found', 'Removed with Remove Duplicates (Data tab) / drop_duplicates()'],
    ['Attrition casing', "~2% of Attrition values were lowercase ('yes'/'no')", 'Standardized with PROPER()/str.capitalize()'],
    ['OverTime whitespace', "~1.5% of OverTime values had leading/trailing spaces", 'Cleaned with TRIM()'],
    ['Department casing', "~1% of Department values were uppercase", 'Standardized with PROPER()/str.title()'],
    ['Missing MonthlyIncome', '18 rows had blank MonthlyIncome', 'Imputed with column median'],
    ['Missing JobSatisfaction', '12 rows had blank JobSatisfaction', 'Imputed with column mode'],
    ['Missing YearsAtCompany', '9 rows had blank YearsAtCompany', 'Imputed with column median'],
    ['Constant columns', 'EmployeeCount, StandardHours, Over18 had a single value for all rows', 'Dropped — no analytical value'],
    ['Derived columns added', 'Raw data only had numeric satisfaction/tenure codes', 'Added TenureBand, IncomeBand, JobSatisfactionLabel, WorkLifeBalanceLabel, AttritionFlag for easier grouping'],
], columns=['Issue', 'Description', 'Fix Applied'])
write_df(ws_log, log_data)
ws_log.column_dimensions['B'].width = 50
ws_log.column_dimensions['C'].width = 55

# ---------------- Clean_Data ----------------
ws_clean = wb.create_sheet('Clean_Data')
write_df(ws_clean, clean)
ws_clean.freeze_panes = 'A2'
last_row = len(clean) + 1
last_col = len(clean.columns)
last_col_letter = get_column_letter(last_col)

tbl = Table(displayName='CleanData', ref=f'A1:{last_col_letter}{last_row}')
tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium11', showRowStripes=True)
ws_clean.add_table(tbl)

col_idx = {c: i + 1 for i, c in enumerate(clean.columns)}
def col(name):
    return get_column_letter(col_idx[name])

rng_attrition = f'Clean_Data!{col("Attrition")}2:{col("Attrition")}{last_row}'
rng_emp = f'Clean_Data!{col("EmployeeNumber")}2:{col("EmployeeNumber")}{last_row}'
rng_income = f'Clean_Data!{col("MonthlyIncome")}2:{col("MonthlyIncome")}{last_row}'
rng_overtime = f'Clean_Data!{col("OverTime")}2:{col("OverTime")}{last_row}'
rng_dept = f'Clean_Data!{col("Department")}2:{col("Department")}{last_row}'
rng_satisfaction = f'Clean_Data!{col("JobSatisfaction")}2:{col("JobSatisfaction")}{last_row}'
rng_tenure = f'Clean_Data!{col("YearsAtCompany")}2:{col("YearsAtCompany")}{last_row}'

# ---------------- KPI_Summary ----------------
ws_kpi = wb.create_sheet('KPI_Summary')
ws_kpi['A1'] = 'KPI Summary (live formulas — change Clean_Data and these recalculate)'
ws_kpi['A1'].font = TITLE_FONT
ws_kpi.merge_cells('A1:C1')

kpi_rows = [
    ('Metric', 'Value', 'Formula Used'),
    ('Total Employees', f'=COUNTA({rng_emp})', 'COUNTA'),
    ('Employees Who Left', f'=COUNTIF({rng_attrition},"Yes")', 'COUNTIF'),
    ('Attrition Rate (%)', f'=B5/B4*100', 'Left / Total'),
    ('Avg Monthly Income (All)', f'=AVERAGE({rng_income})', 'AVERAGE'),
    ('Avg Income — Attrition Yes', f'=AVERAGEIF({rng_attrition},"Yes",{rng_income})', 'AVERAGEIF'),
    ('Avg Income — Attrition No', f'=AVERAGEIF({rng_attrition},"No",{rng_income})', 'AVERAGEIF'),
    ('OverTime Employees', f'=COUNTIF({rng_overtime},"Yes")', 'COUNTIF'),
    ('Attrition Rate — OverTime Yes (%)', f'=COUNTIFS({rng_overtime},"Yes",{rng_attrition},"Yes")/COUNTIF({rng_overtime},"Yes")*100', 'COUNTIFS / COUNTIF'),
    ('Attrition Rate — OverTime No (%)', f'=COUNTIFS({rng_overtime},"No",{rng_attrition},"Yes")/COUNTIF({rng_overtime},"No")*100', 'COUNTIFS / COUNTIF'),
    ('Avg Tenure (Years)', f'=AVERAGE({rng_tenure})', 'AVERAGE'),
    ('Avg Job Satisfaction (1-4)', f'=AVERAGE({rng_satisfaction})', 'AVERAGE'),
]
for i, row_vals in enumerate(kpi_rows, start=3):
    for j, val in enumerate(row_vals, start=1):
        cell = ws_kpi.cell(row=i, column=j, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        if j == 2 and i > 3:
            cell.number_format = '#,##0.00'
style_header(ws_kpi, 3, row=3)
ws_kpi.column_dimensions['A'].width = 32
ws_kpi.column_dimensions['B'].width = 20
ws_kpi.column_dimensions['C'].width = 38

# ---------------- Pivot_Summary (Department x Attrition Rate) ----------------
ws_piv = wb.create_sheet('Pivot_Summary')
ws_piv['A1'] = 'Attrition Rate by Department (build as a real PivotTable in Excel: Insert > PivotTable on Clean_Data)'
ws_piv['A1'].font = TITLE_FONT
ws_piv.merge_cells('A1:D1')

departments = sorted(clean['Department'].dropna().unique().tolist())
piv_header = ['Department', 'Total Employees', 'Attrition Count', 'Attrition Rate (%)']
for j, h in enumerate(piv_header, start=1):
    ws_piv.cell(row=3, column=j, value=h)
style_header(ws_piv, 4, row=3)

for i, dept in enumerate(departments, start=4):
    ws_piv.cell(row=i, column=1, value=dept).font = BODY_FONT
    total_cell = ws_piv.cell(row=i, column=2, value=f'=COUNTIF({rng_dept},A{i})')
    total_cell.font = BODY_FONT
    attr_cell = ws_piv.cell(row=i, column=3, value=f'=COUNTIFS({rng_dept},A{i},{rng_attrition},"Yes")')
    attr_cell.font = BODY_FONT
    rate_cell = ws_piv.cell(row=i, column=4, value=f'=C{i}/B{i}*100')
    rate_cell.font = BODY_FONT
    rate_cell.number_format = '0.00'
    for c in range(1, 5):
        ws_piv.cell(row=i, column=c).border = BORDER

ws_piv.column_dimensions['A'].width = 26
ws_piv.column_dimensions['B'].width = 16
ws_piv.column_dimensions['C'].width = 16
ws_piv.column_dimensions['D'].width = 18

last_dept_row = 3 + len(departments)
chart = BarChart()
chart.title = 'Attrition Rate by Department (%)'
chart.y_axis.title = 'Attrition Rate (%)'
chart.x_axis.title = 'Department'
chart.style = 10
data_ref = Reference(ws_piv, min_col=4, min_row=3, max_row=last_dept_row)
cats_ref = Reference(ws_piv, min_col=1, min_row=4, max_row=last_dept_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 18
chart.height = 10
ws_piv.add_chart(chart, 'F3')

wb.save('/home/claude/hr-attrition/excel/HR_Attrition_Analysis.xlsx')
print('Workbook saved with KPI summary, pivot-style table, and chart')
